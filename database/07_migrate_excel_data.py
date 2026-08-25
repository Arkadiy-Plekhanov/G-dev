"""
Миграция Клауд_финал_Система_развития_личности.xlsx -> PostgreSQL (schema_v2_multitenant.sql)

Логика соответствует Behavioral Equivalence Guarantee из Canonical Spec:
- переносим только фактически введённые пользователем данные (не служебные/формульные колонки)
- явно помеченную демо-строку (Действие Д-0004, "намеренно без цели") НЕ переносим -
  это была живая иллюстрация работы флага в Excel, а не реальная история пользователя
- статусы/приоритеты/группы/контексты маппятся на коды option_lists/quality_groups/action_contexts
- человекочитаемые ID (Ц-0001 и т.п.) сохраняются в legacy_code
"""
import openpyxl
import psycopg2
import bcrypt
import uuid
import sys
import os

XLSX = "/mnt/user-data/uploads/Клауд_финал_Система_развития_личности.xlsx"
DSN = os.environ.get("SEED_DSN", "host=127.0.0.1 dbname=selfdev user=app_writer password=change_me_in_production")

# --- маппинг Excel label -> option_lists.code -------------------------------
GOAL_STATUS = {"Идея": "idea", "Активна": "active", "Приостановлена": "paused",
               "Достигнута": "achieved", "Отменена": "cancelled"}
PRIORITY = {"P1 — Критический": "p1_critical", "P2 — Высокий": "p2_high",
            "P3 — Обычный": "p3_normal", "Фоновый": "background"}
ACTION_STATUS = {"Запланировано": "planned", "Завершено": "done", "Отменено": "cancelled"}
QUALITY_DEV_STATUS = {"Не развито": "undeveloped", "Формируется": "forming",
                       "Устойчиво": "stable", "Закреплено": "anchored"}
QUALITY_FOCUS = {"Текущий фокус": "current_focus", "Поддержание": "maintenance",
                  "Фоновое": "background", "Не в фокусе": "not_in_focus"}
GROUP_CODE = {"Интеллект": "intellect", "Воля": "will", "Самообладание": "self_control",
              "Нравственность": "morality", "Отношения": "relationships", "Лидерство": "leadership",
              "Ответственность": "responsibility", "Обучение": "learning", "Сознательность": "awareness"}
CONTEXT_CODE = {"Переговоры": "negotiation", "Конфликт": "conflict", "Работа": "work",
                 "Публичное выступление": "public_speaking", "Обучение": "learning",
                 "Отношения": "relationships", "Самостоятельная работа": "solo_work",
                 "Общественная среда": "community", "Здоровье/быт": "health_daily", "Другое": "other"}
CYCLE_STATUS = {"Планируется": "planned", "Активен": "active", "Завершён": "done"}
REFLECTION_TYPE = {"Ежедневная": "daily", "Еженедельная": "weekly",
                    "Ежемесячная": "monthly", "По циклу": "cycle"}

EXCLUDE_ACTION_LEGACY_CODES = {"Д-0004"}  # намеренная демо-строка "без цели"


def new_id():
    return str(uuid.uuid4())


def rows(ws):
    headers = [(c.value or "").replace("\n", " ").strip() if isinstance(c.value, str) else c.value
               for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in (None, ""):
            continue
        yield dict(zip(headers, row))


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    conn = psycopg2.connect(DSN)
    cur = conn.cursor()

    # RLS: работаем от лица создаваемого пользователя -> сначала создаём его role-агностично,
    # затем на всю транзакцию переключаем app.current_user_id.
    user_id = new_id()
    pw_hash = bcrypt.hashpw(b"change-me-after-first-login", bcrypt.gensalt()).decode()
    cur.execute(
        "INSERT INTO users (id, email, password_hash, display_name) VALUES (%s,%s,%s,%s)",
        (user_id, "richard@selfdev.local", pw_hash, "Richard"),
    )
    conn.commit()

    cur.execute("SELECT set_config('app.current_user_id', %s, false)", (user_id,))

    # --- Цели -----------------------------------------------------------
    goal_id_map = {}
    ws = wb["Цели"]
    goal_rows = list(rows(ws))
    for r in goal_rows:
        goal_id_map[r["ID"]] = new_id()
    for r in goal_rows:
        cur.execute(
            """INSERT INTO goals (id, user_id, legacy_code, parent_id, name, description,
                                   status_code, priority_code, start_date, target_date, progress_pct)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (goal_id_map[r["ID"]], user_id, r["ID"],
             goal_id_map.get(r["ID родителя"]) if r["ID родителя"] else None,
             r["Название"], r["Описание"],
             GOAL_STATUS[r["Статус"]], PRIORITY[r["Приоритет"]],
             r["Дата начала"], r["Целевая дата"],
             round((r["Прогресс, %"] or 0) * 100, 2)),
        )
    conn.commit()
    print(f"Цели: {len(goal_rows)}")

    # --- Качества ---------------------------------------------------------
    group_code_to_id = {}
    cur.execute("SELECT id, code FROM quality_groups")
    for gid, code in cur.fetchall():
        group_code_to_id[code] = gid

    quality_id_map = {}
    ws = wb["Качества"]
    quality_rows = list(rows(ws))
    for r in quality_rows:
        quality_id_map[r["ID"]] = new_id()
    for r in quality_rows:
        cur.execute(
            """INSERT INTO qualities (id, user_id, legacy_code, name, definition, group_id,
                                       dev_priority_code, focus_code, dev_status_code,
                                       current_level, last_reviewed_at, next_review_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (quality_id_map[r["ID"]], user_id, r["ID"], r["Название"], r["Определение"],
             group_code_to_id.get(GROUP_CODE.get(r["Группа"])),
             PRIORITY[r["Приоритет развития"]], QUALITY_FOCUS[r["Фокус развития"]],
             QUALITY_DEV_STATUS[r["Статус развития"]],
             r["Текущий уровень (РУЧНОЙ, 0-4)"],
             r["Дата последнего пересмотра"], r["Дата следующего пересмотра"]),
        )
    conn.commit()
    print(f"Качества: {len(quality_rows)}")

    # --- Действия -----------------------------------------------------
    context_code_to_id = {}
    cur.execute("SELECT id, code FROM action_contexts")
    for cid, code in cur.fetchall():
        context_code_to_id[code] = cid

    action_id_map = {}
    ws = wb["Действия"]
    action_rows = [r for r in rows(ws) if r["ID"] not in EXCLUDE_ACTION_LEGACY_CODES]
    excluded = [r for r in rows(ws) if r["ID"] in EXCLUDE_ACTION_LEGACY_CODES]
    for r in action_rows:
        action_id_map[r["ID"]] = new_id()
    for r in action_rows:
        cur.execute(
            """INSERT INTO actions (id, user_id, legacy_code, goal_id, name, occurred_at,
                                     description, context_id, result, note, status_code)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (action_id_map[r["ID"]], user_id, r["ID"],
             goal_id_map.get(r["ID цели"]) if r["ID цели"] else None,
             r["Действие"], r["Дата"], r["Описание"],
             context_code_to_id.get(CONTEXT_CODE.get(r["Контекст"])),
             r["Результат"], r["Примечание"], ACTION_STATUS[r["Статус"]]),
        )
    conn.commit()
    print(f"Действия: {len(action_rows)} перенесено, {len(excluded)} исключено (демо-строка): "
          f"{[r['ID'] for r in excluded]}")

    # --- Качества в действиях -> quality_expressions -----------------------
    ws = wb["Качества в действиях"]
    expr_rows = [r for r in rows(ws) if r["ID действия"] not in EXCLUDE_ACTION_LEGACY_CODES]
    for r in expr_rows:
        is_relevant = r["Релевантность"] == "Релевантно"
        cur.execute(
            """INSERT INTO quality_expressions (user_id, legacy_code, action_id, quality_id, is_relevant, score, comment)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (user_id, r["ID"], action_id_map[r["ID действия"]], quality_id_map[r["ID качества"]],
             is_relevant, r["Оценка (0–4)"] if is_relevant else None, r["Комментарий"]),
        )
    conn.commit()
    print(f"Качества в действиях: {len(expr_rows)}")

    # --- Циклы развития + junction ------------------------------------
    ws = wb["Циклы развития"]
    cycle_rows = list(rows(ws))
    cycle_id_map = {}
    for r in cycle_rows:
        cid = new_id()
        cycle_id_map[r["ID"]] = cid
        cur.execute(
            """INSERT INTO development_cycles (id, user_id, legacy_code, name, start_date, end_date,
                                                 status_code, description, summary)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (cid, user_id, r["ID"], r["Название"], r["Дата начала"], r["Дата окончания"],
             CYCLE_STATUS[r["Статус"]], r["Описание"], r["Итоги"]),
        )
        for goal_legacy in [x.strip() for x in (r["Цели (ID через запятую)"] or "").split(",") if x.strip()]:
            if goal_legacy in goal_id_map:
                cur.execute("INSERT INTO cycle_goals (user_id, cycle_id, goal_id) VALUES (%s,%s,%s)",
                            (user_id, cid, goal_id_map[goal_legacy]))
        for q_legacy in [x.strip() for x in (r["Качества в фокусе (ID через запятую)"] or "").split(",") if x.strip()]:
            if q_legacy in quality_id_map:
                cur.execute("INSERT INTO cycle_qualities (user_id, cycle_id, quality_id) VALUES (%s,%s,%s)",
                            (user_id, cid, quality_id_map[q_legacy]))
    conn.commit()
    print(f"Циклы развития: {len(cycle_rows)}")

    # --- Рефлексия --------------------------------------------------------
    ws = wb["Рефлексия"]
    refl_rows = list(rows(ws))
    for r in refl_rows:
        cur.execute(
            """INSERT INTO reflections (user_id, legacy_code, occurred_at, reflection_type_code,
                                         goal_id, cycle_id, what_worked, what_did_not_work,
                                         qualities_observed_raw, insight, what_to_change,
                                         qualities_needing_attention_raw, what_stuck, next_cycle_change)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (user_id, r["ID"], r["Дата"], REFLECTION_TYPE[r["Тип"]],
             goal_id_map.get(r["ID цели"]) if r["ID цели"] else None,
             cycle_id_map.get(r["ID цикла"]) if r["ID цикла"] else None,
             r["Что получилось?"], r["Что не получилось?"], r["Какие качества проявились?"],
             r["Что я понял?"], r["Что изменить?"], r["Какие качества требуют внимания?"],
             r["Что закрепилось?"], r["Что изменить в следующем цикле?"]),
        )
    conn.commit()
    print(f"Рефлексия: {len(refl_rows)}")

    print(f"\nГотово. user_id = {user_id}")
    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
